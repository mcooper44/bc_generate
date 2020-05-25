#!/usr/bin/python3.6
# A utility for printing attendance intake sheets for our
# good good friend L2F and speeding up the intake process
# for people who have pre-registered

import os
import sys

import barcode
from barcode.writer import ImageWriter

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Border 

from file_iface import Menu

WRITER_OPTIONS = {'font_size':4,
                  'text_distance':1,
                  'module_height': 3,
                  'quiet_zone': 1.5}

TOP = Border(top=border.top,
             border_style='double')
BOTTOM= Border(bottom=border.bottom,
               border_style='double')
TRCORNER= Border(top=border.top,
                 right=border.right, 
                 border_style='double')
TLCORNER= Border(top=border.top,
                 left=border.left
                 border_style='double')
BRCORNER= Border(bottom=border.bottom,
                 right=border.right,
                 border_style='double')
BLCORNER= Border(bottom=border.bottom,
                 left=border.left
                 border_style='double')

INCREMENT = 2 # how many spaces to place between barcode lines
ID_INCREMENT = 7 # how many spaces to place between id cards
HEADERS = ['Client ID', 'Client First Name', 'Household Size']
CODE = 'code128'
CODE128 = barcode.get_barcode_class(CODE)

trans_table = str.maketrans({'0': '', '1': '', '2': '', '3': '', '4': '',
                             '5': '', '6': '', '7': '', '8': '', '9': ''})

SOURCE = 'source_files/'
DESTINATION = 'bar_codes/'
NAME = 'test_source.xlsx'


def create_bc(val_str, CODECLASS):
    '''
    creates a bar code image and returns the str name of the png file
    it creates in the DESTINATION folder
    :param: val_str = the string to convert into barcode
    in our case a file ID number
    :param: CODECLASS is the barcode standard class 
    stored in the CODE128 variable
    '''
    code = CODECLASS(val_str, writer=ImageWriter())
    result = code.save(f'{DESTINATION}{val_str}', options=WRITER_OPTIONS)
    if not result:
        raise ValueError(f'Could not save barcode for {val_str}')
    else:
        return result

def add_image(code_file, cell_str, ws_handl):
    '''
    adds a barcode image to the ws_handl at location
    cell_str
    param: code_file = a path to a file, or the image file
    param: cell_str = the location to insest the image i.e. A3
    param: ws_handle = the worksheet handler initialized earlier
    '''
    img = openpyxl.drawing.image.Image(code_file)
    img.anchor = cell_str # i.e. B2
    ws_handl.add_image(img)


def put_code(code_file, cell_str, ws_handl, file_info, loop_num):
    '''
    This is for generating a sheet of file info and barcodes
    takes a image file name, a cell location
    and a worksheet and saves the image to the worksheet at
    the cell location specified
    param: code_file = image to append
    param: cell_str = cell location to paste in file
    param: ws_handl = work sheet handler to interact with
    param: file_info = list containing values to append to file
    param: loop_num = the number of times a line has been written
    in line with image
    '''
    try:
        add_image(code_file, cell_str, ws_handl)
        
        if INCREMENT > 1 and loop_num > 1:
            for x in range(INCREMENT-1):
                ws_handl.append(['','',''])
        ws_handl.append(file_info) #  123456, john, smith
    except:
        raise ValueError(f'Could not set image file {code_file} at {cell_str}')

def put_id_card(code_file, cell_str, ws_handl, name_string, name_index):

    try:
        add_image(code_file, cell_str, ws_handl)
        ws_handl[name_index] = name_string
    except:
        raise ValueError(f'Could not set image {code_file}\nor name {name_string} at {name_index}')

def fnd_col_lttr(cell):
    '''
    param: cell = the openpyxl cell class object
    openpyl repr for a cell is in the format of <Cell 'Sheet1'.A1>
    this function extracts the Column Letter A from that
    str and returns it or raises and error
    '''
    
    label = str(cell).split('.')[1].strip('>').translate(trans_table)
    
    if all(l.isalpha() for l in label):
        return label
    else:
        raise ValueError(f'cell {cell} is invalid could not find a column label')

def fnd_sub_str(rng, sub_string):
    '''
    param: rng = the tuple structure returned by openpyxl for a row.  it 
    iterates through the tuple structure of a row looking for the substring
    when it finds it it returns the column where the sub string is the header
    or None if it cannot find it or  raises an error if param rng is not valid
    '''
    try:
        for item in rng:
            if sub_string in item.value:
                return fnd_col_lttr(item)
    except Exception:
        raise Exception(f'could not find a column with {sub_string}')

def file_set(target_dir='bar_codes/'):
    '''
    returns a list of image files in the directory as a set
    this is for evaluating if a barcode image has already been 
    generated or if one needs to be created
    '''    
    return set(x for x in os.listdir(target_dir) if x.endswith('.png'))


def return_bars(cell_val, bar_code_files):
    '''
    looks for a file name cell_val.png in the previously generated
    bar code files, if it does not exist, it generates a barcode
    and returns it otherwise, it returns the path to the file
    so openpyxl can insert the file to the sheet
    '''
    bars = f'{DESTINATION}{cell_val}.png' # path/1111111.png
    if bars.split('/')[1] not in bar_code_files:        
        bars = create_bc(cell_val, CODE128) # create barcode for File ID
    return bars

def write_code_sheet(cell_val, cell_index, ws_bc, info_line,
                     LOOP,bar_code_files): 
    '''
    writes a info_line at cell_index into ws_bc 
    and appends an image by looking for a file
    in bar_code_files named after cell_val and if it doesn't find it calling return bars
    to generate one using cell_val and then insert the barcode into the worksheet
    '''

    image_index = cell_index
    if cell_index >2:
        image_index = LOOP * INCREMENT    
    bars = return_bars(cell_val, bar_code_files)
    put_code(bars, f'D{image_index}', ws_bc, info_line,LOOP)
    ws_bc.row_dimensions[image_index].height = 65

def write_id_cards(LOOP, id_switch, cell_val, full_nm, id_dex, bar_code_files, ws_id):
    
    bars = return_bars(cell_val, bar_code_files) # path or image file

    if not id_switch:
        name1_l = f'A{id_dex}'#1
        card1_l = f'A{id_dex+1}'#2
        #print(f'{cell_val} id_dex {id_dex} name {name1_l}\n') 
        put_id_card(bars, card1_l, ws_id, full_nm, name1_l)
    else:
        name2_l =f'F{id_dex}'#1
        card2_l = f'F{id_dex+1}'#2
        #print(f'{cell_val} id_dex {id_dex} name {name2_l}\n') 
        put_id_card(bars, card2_l, ws_id, full_nm, name2_l)

def connect_xl_file(fname,codes=True,cards=False):
    '''
    opens a connection to an excel file (fname)
    and returns the workbook and ws handlers as well
    as a tuple of key variables:
    cell_index = the row number to start reading file numbers from
    default is wired to be line 2 of the file
    col = the label for the column that the file id's are on i.e F 
    fname, lname = the first/last name column letters i.e A, B

    '''
    wb = load_workbook(fname)
    ws = wb.active
    ws_bc = None
    cell_index = 2
    
    if codes:
        ws_bc = wb.create_sheet('barcodes')
        ws_bc.append(['File ID', 'F. Name','L. Name', 'Barcode'])
        ws_bc.column_dimensions['D'].width = 42
    if cards:
        ws_bc = wb.create_sheet('ID_Cards')

    col = fnd_sub_str(ws[1], 'Client ID')# i.e. A
    fname = fnd_sub_str(ws[1], 'Client First Name')
    lname = fnd_sub_str(ws[1], 'Client Last Name')

    return wb, ws, ws_bc, (cell_index, col, fname, lname)

def handle_xl_file(filename,bcsheet=True,idcards=False):
    '''
    opens the active worksheet and looks for the 
    following headings: Client ID, Client First Name, Client Last Name

    when it finds them it iterates down the sheet, pulling out
    those cells and appending them to a new worksheet called barcodes
    with the barcode images in the 4th column
    '''
    LOOP = 1
    id_dex = 1
    id_switch = False
    wb, ws, ws_bc, dexs = connect_xl_file(filename,codes=bcsheet,cards=idcards)
    cell_index, col, f_name, lname  = dexs # 2, A, B, C
    bar_code_files = file_set() # bar code image files
    
    if idcards: LOOP = 0

    for n in range(len(ws[col])):
        n_l = f'{col}{cell_index}' # i.e. A2
        cell_val = str(ws[n_l].value) # File ID string at A2
        f_l = f'{f_name}{cell_index}'# B2
        f_val = str(ws[f_l].value) # First name stored at B2
        l_l = f'{lname}{cell_index}'
        l_val = str(ws[l_l].value)
        info_line = [cell_val, f_val, l_val] # [12345, John, Smith]

        if cell_val != 'None':
            
            if bcsheet:
                # number to be encoded, row number, worhseet, [info line],
                # iteration, location of code files
                write_code_sheet(cell_val, cell_index, ws_bc, info_line,
                                 LOOP,bar_code_files)
            if idcards:
                full_nm = f'{f_val} {l_val}'
                

                write_id_cards(LOOP,id_switch, cell_val, full_nm,id_dex,
                               bar_code_files, ws_bc)
                id_switch = id_switch ^ True

            LOOP += 1
            if LOOP % 2 == 0:
                id_dex += ID_INCREMENT
            if LOOP % 12 == 0:
                id_dex += 2
        cell_index += 1 # next time through we'll operate on A3
    wb.save(filename)
    
def main():
    print('Choose source csv file')
    menu = Menu(base_path=SOURCE)
    menu.get_file_list()
    target = menu.handle_input(menu.prompt_input('files'))
    
    print('Please confirm your choice')
    confirm = input(f'please choose...\n1. Confirm {target}\n2. Exit\n')
    if confirm == '1':    
        operation = input(f'Select \n1. Barcode sheet\n2. ID Cards\n3. Exit\n')
        if operation == '1':
            handle_xl_file(target)
        if operation == '2':
            handle_xl_file(target, bcsheet=False, idcards=True)
        else:
            print('exiting...')
            sys.exit(1)
    else:
        print('exiting...')
        sys.exit(1)
    print('we are done!')

if __name__ == '__main__':
    main()             
